import os
import yaml
from openpyxl import Workbook
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, colors


def get_path(filename):
    return os.path.join(os.path.dirname(__file__), filename)


def read_config():
    with open("Config.yaml", 'r') as yml_file:
        cfg = yaml.load(yml_file, yaml.SafeLoader)
    return cfg


def create_excel_report(save_folder, attributes, model_data):
    print("\nCreating Report..")

    columns = []
    widths = []

    wb = Workbook()
    report = wb.active
    report.title = datetime.now().strftime('Report')

    for attribute in attributes:
        columns.append(attribute)
        widths.append(40)

    # Styles
    # alignment_titles = Alignment(horizontal='left', vertical='left', wrap_text=True)
    fill_titles = PatternFill("solid", fgColor="ad0f5b")
    font_titles = Font(color=colors.WHITE, bold=True, size=12)

    # Insert all column titles
    for i in range(len(columns)):
        cell = report.cell(column=(i + 1), row=1)
        cell.font = font_titles
        cell.fill = fill_titles
        # cell.alignment = alignment_titles
        cell.value = columns[i]
        report.column_dimensions[get_column_letter(i + 1)].width = widths[i]
        # report.row_dimensions[2].height = 30

    row = 1
    for modelAttributes in model_data:
        row += 1
        column = 1
        for iAttr in range(len(modelAttributes)):
            report.cell(column=column, row=row, value=modelAttributes[iAttr])
            column += 1

    print('Report Location : ' + save_folder)

    if not os.path.exists(save_folder):
        os.makedirs(save_folder, exist_ok=True)

    output_file_name = 'SPM_Report_' + '%Y.%m.%d_%H.%M.%S.xlsx'
    save = datetime.now().strftime(output_file_name)
    file_name = os.path.join(save_folder, save)
    wb.save(file_name)

    return file_name
